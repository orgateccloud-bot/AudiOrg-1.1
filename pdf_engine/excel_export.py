"""Exportação Excel de NFAs — fonte única da verdade para todas as saídas Excel.

Substitui a duplicação entre app.py (_exportar_excel_core) e extrair_nfa.py (exportar_excel).
"""

from __future__ import annotations

from collections import defaultdict
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from nfa_extractor.domain.constants import CABECALHOS_EXCEL, CORES, LARGURAS_EXCEL

if TYPE_CHECKING:
    from nfa_extractor.domain.extractor import NFA


def _cel_style(
    ws,
    row: int,
    col: int,
    value,
    bold: bool = False,
    bg: str | None = None,
    fg: str = '000000',
    align: str = 'left',
    num_fmt: str | None = None,
    size: int = 9,
):
    """Aplica estilo padronizado a uma célula do Excel."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Calibri', bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
    thin = Side(style='thin', color=CORES['BORDER'])
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _cabecalho_linha(ws, row: int, valores: list, larguras: list) -> None:
    """Renderiza uma linha de cabeçalho com cores e largura de colunas."""
    for j, (v, w) in enumerate(zip(valores, larguras), 1):
        _cel_style(ws, row, j, v, bold=True, bg='1E3A8A', fg='FFFFFF', align='center')
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row].height = 20


def _configurar_aba(ws) -> None:
    """Aplica configurações padrão em uma aba: freeze + sem gridlines."""
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False


def _cor_linha(numero_linha: int) -> str:
    """Retorna a cor de fundo alternado para linhas de dados."""
    return 'F3F4F6' if numero_linha % 2 == 0 else 'FFFFFF'


def _aba_notas(wb, notas: list[NFA]) -> None:
    """Aba 1 — Resumo por nota fiscal."""
    ws = wb.active
    ws.title = 'Notas Fiscais'
    _configurar_aba(ws)
    _cabecalho_linha(ws, 1, CABECALHOS_EXCEL['notas'], LARGURAS_EXCEL['notas'])

    colunas_numericas = {9, 10}
    for i, n in enumerate(notas, 2):
        bg = _cor_linha(i)
        valores = [
            n.numero, n.emissao, n.natureza, n.local_emissao,
            n.destinatario.nome, n.destinatario.cpf_cnpj, n.destinatario.municipio,
            n.transportador.nome, n.quantidade_total, n.valor_total,
            n.chave_acesso,
        ]
        for j, v in enumerate(valores, 1):
            fmt = '#,##0.00' if j in colunas_numericas else None
            _cel_style(ws, i, j, v, bg=bg, fg='000000',
                       align='right' if j in colunas_numericas else 'left',
                       num_fmt=fmt)
        ws.row_dimensions[i].height = 18


def _aba_itens(wb, notas: list[NFA]) -> None:
    """Aba 2 — Itens detalhados (um produto por linha)."""
    ws = wb.create_sheet('Itens Detalhados')
    _configurar_aba(ws)
    _cabecalho_linha(ws, 1, CABECALHOS_EXCEL['itens'], LARGURAS_EXCEL['itens'])

    colunas_numericas = {9, 10, 11, 12}
    row = 2
    for n in notas:
        for prod in n.produtos:
            bg = _cor_linha(row)
            valores = [
                n.numero, n.emissao, n.natureza,
                n.destinatario.nome, n.destinatario.cpf_cnpj, n.destinatario.municipio,
                prod.codigo, prod.descricao,
                prod.quantidade, prod.vlr_unitario, prod.vlr_icms, prod.vlr_total,
            ]
            for j, v in enumerate(valores, 1):
                fmt = '#,##0.00' if j in colunas_numericas else None
                _cel_style(ws, row, j, v, bg=bg, fg='000000',
                           align='right' if j in colunas_numericas else 'left',
                           num_fmt=fmt)
            ws.row_dimensions[row].height = 18
            row += 1


def _aba_destinatarios(wb, notas: list[NFA]) -> None:
    """Aba 3 — Resumo por destinatário com linha de totais."""
    ws = wb.create_sheet('Por Destinatário')
    _configurar_aba(ws)
    _cabecalho_linha(ws, 1, CABECALHOS_EXCEL['dest'], LARGURAS_EXCEL['dest'])

    # Agrupamento por destinatário
    agrup: dict[str, dict] = defaultdict(lambda: {
        'nome': '', 'cpf_cnpj': '', 'municipio': '',
        'notas': 0, 'cabecas': 0.0, 'valor': 0.0,
    })
    for n in notas:
        k = n.destinatario.cpf_cnpj or n.destinatario.nome
        agrup[k]['nome']      = n.destinatario.nome
        agrup[k]['cpf_cnpj']  = n.destinatario.cpf_cnpj
        agrup[k]['municipio'] = n.destinatario.municipio
        agrup[k]['notas']    += 1
        agrup[k]['cabecas']  += n.quantidade_total
        agrup[k]['valor']    += n.valor_total

    ordenados = sorted(agrup.values(), key=lambda x: x['valor'], reverse=True)
    colunas_numericas = {4, 5, 6, 7}
    for i, d in enumerate(ordenados, 2):
        bg = _cor_linha(i)
        ticket = d['valor'] / d['notas'] if d['notas'] else 0.0
        valores = [d['nome'], d['cpf_cnpj'], d['municipio'], d['notas'], d['cabecas'], d['valor'], ticket]
        for j, v in enumerate(valores, 1):
            fmt = '#,##0.00' if j in colunas_numericas else None
            _cel_style(ws, i, j, v, bg=bg, fg='000000',
                       align='right' if j in colunas_numericas else 'left',
                       num_fmt=fmt)
        ws.row_dimensions[i].height = 18

    # Linha de totais
    tot_row = len(ordenados) + 2
    for j in range(1, len(CABECALHOS_EXCEL['dest']) + 1):
        _cel_style(ws, tot_row, j, '', bold=True, bg='1E3A8A', fg='FFFFFF')
    _cel_style(ws, tot_row, 1, 'TOTAL', bold=True, bg='1E3A8A', fg=CORES['CYAN'], size=10)
    _cel_style(ws, tot_row, 4, sum(d['notas']   for d in ordenados),
               bold=True, bg='1E3A8A', fg=CORES['GREEN'], align='right')
    _cel_style(ws, tot_row, 5, sum(d['cabecas'] for d in ordenados),
               bold=True, bg='1E3A8A', fg=CORES['GREEN'], align='right', num_fmt='#,##0.00')
    _cel_style(ws, tot_row, 6, sum(d['valor']   for d in ordenados),
               bold=True, bg='1E3A8A', fg=CORES['GREEN'], align='right', num_fmt='#,##0.00')


def _aba_mensal(wb, notas: list[NFA], nome_contribuinte: str = "") -> None:
    """Aba 4 — Evolução mensal do faturamento."""
    from nfa_extractor.domain.extractor import resumo_geral

    ws = wb.create_sheet('Evolução Mensal')
    _configurar_aba(ws)
    _cabecalho_linha(ws, 1, CABECALHOS_EXCEL['mensal'], LARGURAS_EXCEL['mensal'])

    res = resumo_geral(notas, nome_contribuinte=nome_contribuinte)
    colunas_numericas = {3, 4}
    for i, (mes, v) in enumerate(res['por_mes'].items(), 2):
        bg = _cor_linha(i)
        valores = [mes, v['notas'], v['cabecas'], v['valor']]
        for j, val in enumerate(valores, 1):
            fmt = '#,##0.00' if j in colunas_numericas else None
            _cel_style(ws, i, j, val, bg=bg, fg='000000',
                       align='right' if j > 1 else 'left', num_fmt=fmt)
        ws.row_dimensions[i].height = 18


def exportar_excel(notas: list[NFA], saida: str, nome_contribuinte: str = "") -> None:
    """Exporta a lista de NFAs para Excel com 4 abas formatadas.

    Args:
        notas: Lista de objetos NFA extraídos do PDF.
        saida: Caminho do arquivo .xlsx de destino.

    Abas geradas:
        1. Notas Fiscais — uma linha por nota
        2. Itens Detalhados — um produto por linha
        3. Por Destinatário — resumo agrupado por comprador
        4. Evolução Mensal — faturamento mês a mês
    """
    if not notas:
        raise ValueError('Nenhuma nota fiscal para exportar.')

    wb = openpyxl.Workbook()
    _aba_notas(wb, notas)
    _aba_itens(wb, notas)
    _aba_destinatarios(wb, notas)
    _aba_mensal(wb, notas, nome_contribuinte=nome_contribuinte)

    wb.save(saida)
